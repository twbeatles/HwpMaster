"""
Hyperlink Checker Module
HWP 문서 하이퍼링크 검사

Author: HWP Master
"""

import gc
import html as html_lib
import logging
import os
import socket
import fnmatch
import urllib.request
import urllib.error
from concurrent.futures import ThreadPoolExecutor, as_completed
from contextlib import ExitStack
from pathlib import Path
from dataclasses import dataclass, field
from enum import Enum
from threading import Lock
from typing import Optional, Callable, Any
from urllib.parse import urlparse


class LinkStatus(Enum):
    """링크 상태"""

    VALID = "valid"
    BROKEN = "broken"
    TIMEOUT = "timeout"
    UNKNOWN = "unknown"
    LOCAL_MISSING = "local_missing"
    LOCAL_OK = "local_ok"
    SKIPPED = "skipped"


def parse_allowlist(text: str) -> list[str]:
    """
    Parse a comma-separated allowlist string into patterns.

    Examples:
    - "example.com, *.corp.local"
    """
    items: list[str] = []
    for raw in (text or "").split(","):
        s = raw.strip()
        if s:
            items.append(s)
    return items


def host_in_allowlist(host: str, patterns: list[str]) -> bool:
    """
    Best-effort host matching.

    - "*.example.com" matches "a.example.com" and "example.com"
    - "example.com" matches exactly "example.com"
    - patterns with "*" use fnmatch
    """
    host = (host or "").strip().lower()
    if not host:
        return False

    for pat in patterns:
        p = pat.strip().lower()
        if not p:
            continue
        if p.startswith("*."):
            suffix = p[1:]  # ".example.com"
            if host == p[2:] or host.endswith(suffix):
                return True
            continue
        if "*" in p:
            if fnmatch.fnmatch(host, p):
                return True
            continue
        if host == p:
            return True

    return False


@dataclass
class LinkInfo:
    """링크 정보"""

    url: str
    text: str
    page: int
    status: LinkStatus = LinkStatus.UNKNOWN
    error_message: str = ""


@dataclass
class LinkCheckResult:
    """링크 검사 결과"""

    success: bool
    source_path: str
    links: list[LinkInfo] = field(default_factory=list)
    valid_count: int = 0
    broken_count: int = 0
    error_message: Optional[str] = None


class HyperlinkChecker:
    """하이퍼링크 검사기"""

    def __init__(
        self,
        *,
        external_requests_enabled: bool = True,
        timeout_sec: int = 5,
        domain_allowlist: str = "",
        max_concurrency: Optional[int] = None,
        cache_enabled: bool = True,
    ) -> None:
        self._hwp: Any = None
        self._is_initialized = False
        self._logger = logging.getLogger(__name__)
        self._timeout = max(1, int(timeout_sec))
        self._external_requests_enabled = bool(external_requests_enabled)
        self._allowlist_patterns = parse_allowlist(domain_allowlist)
        self._com_stack: Optional[ExitStack] = None

        cpu = os.cpu_count() or 4
        if max_concurrency is None:
            # Conservative default to avoid overloading network/target servers.
            self._max_concurrency = max(1, min(8, cpu))
        else:
            self._max_concurrency = max(1, int(max_concurrency))

        self._cache_enabled = bool(cache_enabled)
        self._url_cache: dict[str, tuple[LinkStatus, str]] = {}
        self._cache_lock = Lock()

    def _ensure_hwp(self) -> None:
        if self._hwp is None:
            try:
                if self._com_stack is None:
                    from ..utils.com_init import com_context

                    self._com_stack = ExitStack()
                    self._com_stack.enter_context(com_context())

                import pyhwpx

                self._hwp = pyhwpx.Hwp(visible=False)
                self._is_initialized = True
            except ImportError:
                if self._com_stack is not None:
                    self._com_stack.close()
                    self._com_stack = None
                raise RuntimeError("pyhwpx가 설치되어 있지 않습니다.")
            except Exception as e:
                if self._com_stack is not None:
                    self._com_stack.close()
                    self._com_stack = None
                raise RuntimeError(f"한글 프로그램 초기화 실패: {e}")

    def _get_hwp(self) -> Any:
        self._ensure_hwp()
        if self._hwp is None:
            raise RuntimeError("한글 인스턴스 초기화 실패")
        return self._hwp

    def close(self) -> None:
        if self._hwp is not None:
            try:
                self._hwp.quit()
            except Exception as e:
                self._logger.warning(f"HWP 종료 중 오류 (무시): {e}")
            finally:
                self._hwp = None
                self._is_initialized = False
                gc.collect()

        if self._com_stack is not None:
            try:
                self._com_stack.close()
            except Exception:
                pass
            finally:
                self._com_stack = None

    def __enter__(self):
        self._ensure_hwp()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()
        return False

    def extract_links(self, source_path: str) -> LinkCheckResult:
        """문서 내 모든 링크 추출"""
        try:
            hwp = self._get_hwp()
            source = Path(source_path)
            if not source.exists():
                return LinkCheckResult(False, source_path, error_message="파일 없음")

            hwp.open(source_path)
            links: list[LinkInfo] = []

            try:
                ctrl: Any = hwp.HeadCtrl
                while ctrl:
                    if ctrl.UserDesc == "하이퍼링크":
                        try:
                            url = str(ctrl.GetSetItem("HyperLink"))
                            text = str(ctrl.GetSetItem("Text") or url)
                            links.append(LinkInfo(url=url, text=text, page=0))
                        except Exception as e:
                            self._logger.warning(f"하이퍼링크 정보 추출 실패: {e}")
                    ctrl = ctrl.Next
            except Exception as e:
                self._logger.warning(f"하이퍼링크 컨트롤 탐색 중 오류: {e}")

            return LinkCheckResult(True, source_path, links)
        except Exception as e:
            self._logger.warning(f"링크 추출 실패: {e}", exc_info=True)
            return LinkCheckResult(False, source_path, error_message=str(e))

    def _check_local_file(self, path: str) -> tuple[LinkStatus, str]:
        """로컬 파일 경로 검사"""
        try:
            if path.startswith("file:///"):
                path = path[8:]
            elif path.startswith("file://"):
                path = path[7:]

            if Path(path).exists():
                return LinkStatus.LOCAL_OK, ""
            return LinkStatus.LOCAL_MISSING, "파일 없음"
        except Exception as e:
            return LinkStatus.UNKNOWN, str(e)

    def check_url(self, url: str) -> tuple[LinkStatus, str]:
        """URL 유효성 검사"""
        if url.startswith("file://") or url.startswith("/") or (len(url) > 1 and url[1] == ":"):
            return self._check_local_file(url)

        parsed = urlparse(url)
        if parsed.scheme not in ("http", "https"):
            return LinkStatus.UNKNOWN, "지원하지 않는 URL 스킴"

        if not self._external_requests_enabled:
            return LinkStatus.SKIPPED, "외부 접속 비활성화"

        host = (parsed.hostname or "").strip()
        if self._allowlist_patterns and not host_in_allowlist(host, self._allowlist_patterns):
            return LinkStatus.SKIPPED, f"allowlist 불일치: {host or '(no host)'}"

        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=self._timeout) as response:
                if response.status < 400:
                    return LinkStatus.VALID, ""
                return LinkStatus.BROKEN, f"HTTP {response.status}"
        except urllib.error.HTTPError as e:
            return LinkStatus.BROKEN, f"HTTP {e.code}"
        except urllib.error.URLError as e:
            if isinstance(getattr(e, "reason", None), socket.timeout):
                return LinkStatus.TIMEOUT, "연결 시간 초과"
            return LinkStatus.BROKEN, str(getattr(e, "reason", e))
        except socket.timeout:
            return LinkStatus.TIMEOUT, "연결 시간 초과"
        except TimeoutError:
            return LinkStatus.TIMEOUT, "연결 시간 초과"
        except Exception as e:
            return LinkStatus.UNKNOWN, str(e)

    def _check_url_cached(self, url: str) -> tuple[LinkStatus, str]:
        if not self._cache_enabled:
            return self.check_url(url)

        with self._cache_lock:
            cached = self._url_cache.get(url)
        if cached is not None:
            return cached

        outcome = self.check_url(url)
        with self._cache_lock:
            self._url_cache[url] = outcome
        return outcome

    def check_links(
        self,
        source_path: str,
        progress_callback: Optional[Callable[[int, int, str], None]] = None,
    ) -> LinkCheckResult:
        """문서 내 모든 링크 추출 및 검사"""
        result = self.extract_links(source_path)
        if not result.success:
            return result

        total = len(result.links)
        if total == 0:
            result.valid_count = 0
            result.broken_count = 0
            return result

        outcomes: list[tuple[LinkStatus, str]] = [(LinkStatus.UNKNOWN, "")] * total

        def _apply_outcomes() -> None:
            valid_count = 0
            broken_count = 0
            for idx, link in enumerate(result.links, start=1):
                status, error = outcomes[idx - 1]
                link.status = status
                link.error_message = error
                if progress_callback:
                    progress_callback(idx, total, link.url[:50])

                if status in (LinkStatus.VALID, LinkStatus.LOCAL_OK):
                    valid_count += 1
                elif status in (LinkStatus.BROKEN, LinkStatus.LOCAL_MISSING, LinkStatus.TIMEOUT):
                    broken_count += 1

            result.valid_count = valid_count
            result.broken_count = broken_count

        if self._max_concurrency <= 1:
            for i, link in enumerate(result.links):
                outcomes[i] = self._check_url_cached(link.url)
            _apply_outcomes()
            return result

        try:
            if self._cache_enabled:
                pending_urls: dict[str, list[int]] = {}

                for idx, link in enumerate(result.links):
                    url = link.url
                    with self._cache_lock:
                        cached = self._url_cache.get(url)
                    if cached is not None:
                        outcomes[idx] = cached
                        continue
                    pending_urls.setdefault(url, []).append(idx)

                if pending_urls:
                    with ThreadPoolExecutor(max_workers=self._max_concurrency) as executor:
                        future_to_url = {
                            executor.submit(self.check_url, url): url for url in pending_urls.keys()
                        }

                        for future in as_completed(future_to_url):
                            url = future_to_url[future]
                            try:
                                outcome = future.result()
                            except Exception as e:
                                outcome = (LinkStatus.UNKNOWN, str(e))

                            with self._cache_lock:
                                self._url_cache[url] = outcome

                            for idx in pending_urls[url]:
                                outcomes[idx] = outcome
            else:
                with ThreadPoolExecutor(max_workers=self._max_concurrency) as executor:
                    future_to_idx = {
                        executor.submit(self.check_url, link.url): idx
                        for idx, link in enumerate(result.links)
                    }

                    for future in as_completed(future_to_idx):
                        idx = future_to_idx[future]
                        try:
                            outcomes[idx] = future.result()
                        except Exception as e:
                            outcomes[idx] = (LinkStatus.UNKNOWN, str(e))

        except Exception as e:
            self._logger.warning(f"병렬 링크 검사 실패, 순차로 폴백: {e}")
            for i, link in enumerate(result.links):
                outcomes[i] = self._check_url_cached(link.url)

        _apply_outcomes()
        return result

    def generate_report(self, result: LinkCheckResult, output_path: str) -> bool:
        """HTML 리포트 생성"""
        try:
            file_name = html_lib.escape(Path(result.source_path).name)
            html_text = f"""<!DOCTYPE html>
<html><head><meta charset=\"utf-8\"><title>링크 검사 리포트</title>
<style>
body {{ font-family: 'Segoe UI', sans-serif; max-width: 900px; margin: 0 auto; padding: 20px; }}
h1 {{ color: #333; }} .summary {{ background: #f5f5f5; padding: 15px; border-radius: 8px; margin: 20px 0; }}
table {{ width: 100%; border-collapse: collapse; }}
th, td {{ padding: 10px; text-align: left; border-bottom: 1px solid #ddd; }}
.valid {{ color: #22c55e; }} .broken {{ color: #ef4444; }} .unknown {{ color: #f59e0b; }}
</style></head><body>
<h1>🔗 링크 검사 리포트</h1>
<div class=\"summary\">
<p><strong>파일:</strong> {file_name}</p>
<p><strong>총 링크:</strong> {len(result.links)}개</p>
<p><strong>유효:</strong> <span class=\"valid\">{result.valid_count}개</span> |
<strong>오류:</strong> <span class=\"broken\">{result.broken_count}개</span></p>
</div>
<table><tr><th>상태</th><th>URL</th><th>텍스트</th><th>오류</th></tr>"""

            for link in result.links:
                if link.status in (LinkStatus.VALID, LinkStatus.LOCAL_OK):
                    status_class = "valid"
                    status_text = "✅"
                elif link.status in (LinkStatus.BROKEN, LinkStatus.LOCAL_MISSING, LinkStatus.TIMEOUT):
                    status_class = "broken"
                    status_text = "❌"
                else:
                    status_class = "unknown"
                    status_text = "❓"

                html_text += (
                    f'<tr><td class="{status_class}">{status_text}</td>'
                    f"<td>{html_lib.escape(link.url)}</td>"
                    f"<td>{html_lib.escape(link.text)}</td>"
                    f"<td>{html_lib.escape(link.error_message)}</td></tr>"
                )

            html_text += "</table></body></html>"

            with open(output_path, "w", encoding="utf-8") as f:
                f.write(html_text)
            return True
        except Exception as e:
            self._logger.warning(f"HTML 리포트 생성 실패: {e}", exc_info=True)
            return False

    def export_links_to_excel(self, links: list[tuple[str, LinkInfo]], output_path: str) -> bool:
        """
        링크 검사 결과를 Excel(.xlsx)로 내보내기.

        Args:
            links: (filename, LinkInfo) 튜플 리스트
            output_path: .xlsx 경로
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            if ws is None:
                return False
            ws.title = "Hyperlinks"

            headers = ["파일명", "상태", "URL", "텍스트", "오류"]
            ws.append(headers)

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="333333")
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(vertical="center")

            valid_fill = PatternFill("solid", fgColor="D1FAE5")
            broken_fill = PatternFill("solid", fgColor="FEE2E2")
            unknown_fill = PatternFill("solid", fgColor="FEF3C7")

            for filename, link in links:
                status = getattr(link.status, "value", str(link.status))
                ws.append([filename, status, link.url, link.text, link.error_message])

                row = ws.max_row
                fill = unknown_fill
                if status in [LinkStatus.VALID.value, LinkStatus.LOCAL_OK.value]:
                    fill = valid_fill
                elif status in [LinkStatus.BROKEN.value, LinkStatus.LOCAL_MISSING.value, LinkStatus.TIMEOUT.value]:
                    fill = broken_fill

                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).fill = fill
                    ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")

            ws.column_dimensions["A"].width = 24
            ws.column_dimensions["B"].width = 14
            ws.column_dimensions["C"].width = 60
            ws.column_dimensions["D"].width = 28
            ws.column_dimensions["E"].width = 40

            wb.save(output_path)
            wb.close()
            return True
        except Exception as e:
            self._logger.warning(f"엑셀 내보내기 실패: {e}")
            return False
