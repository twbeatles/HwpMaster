"""
Excel Handler Module
openpyxl 기반 엑셀 데이터 처리 (Pandas 대체)

Author: HWP Master
"""

import csv
import logging
from pathlib import Path
from typing import Optional, Iterator, Any
from dataclasses import dataclass


_logger = logging.getLogger(__name__)


@dataclass
class ExcelReadResult:
    """엑셀 읽기 결과"""
    success: bool
    data: list[dict[str, Any]]
    headers: list[str]
    row_count: int
    error_message: Optional[str] = None


class ExcelHandler:
    """
    openpyxl 기반 엑셀 데이터 처리 클래스
    Pandas 없이 list[dict] 형태로 데이터 반환
    """
    
    @staticmethod
    def read_excel(
        file_path: str,
        sheet_name: Optional[str] = None,
        header_row: int = 1,
        start_row: Optional[int] = None,
        end_row: Optional[int] = None
    ) -> ExcelReadResult:
        """
        엑셀 파일 읽기
        
        Args:
            file_path: 엑셀 파일 경로
            sheet_name: 시트 이름 (None이면 첫 번째 시트)
            header_row: 헤더 행 번호 (1-indexed)
            start_row: 시작 행 (None이면 헤더 다음 행)
            end_row: 종료 행 (None이면 끝까지)
        
        Returns:
            ExcelReadResult 객체
        """
        wb = None
        try:
            from openpyxl import load_workbook
            
            path = Path(file_path)
            if not path.exists():
                return ExcelReadResult(
                    success=False,
                    data=[],
                    headers=[],
                    row_count=0,
                    error_message=f"파일이 존재하지 않습니다: {file_path}"
                )
            
            # 워크북 로드 (데이터만, 수식 계산하지 않음)
            wb = load_workbook(file_path, data_only=True, read_only=True)
            
            # 시트 선택
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    return ExcelReadResult(
                        success=False,
                        data=[],
                        headers=[],
                        row_count=0,
                        error_message=f"시트가 존재하지 않습니다: {sheet_name}"
                    )
                ws = wb[sheet_name]
            else:
                ws = wb.active

            if ws is None:
                return ExcelReadResult(
                    success=False,
                    data=[],
                    headers=[],
                    row_count=0,
                    error_message="시트를 찾을 수 없습니다."
                )
            
            # 헤더 추출
            headers: list[str] = []
            for cell in ws[header_row]:
                value = cell.value
                if value is not None:
                    headers.append(str(value).strip())
                else:
                    headers.append(f"Column_{cell.column}")
            
            # 데이터 시작 행
            data_start = start_row if start_row else header_row + 1
            
            # 데이터 읽기
            data: list[dict[str, Any]] = []
            row_idx = 0
            
            for row_num, row in enumerate(ws.iter_rows(min_row=data_start, max_row=end_row), start=data_start):
                # 빈 행 감지 (첫 번째 셀이 비어있으면 스킵)
                first_cell_value = row[0].value if row else None
                if first_cell_value is None:
                    continue
                
                row_data: dict[str, Any] = {}
                for idx, cell in enumerate(row):
                    if idx < len(headers):
                        value = cell.value
                        # None은 빈 문자열로
                        row_data[headers[idx]] = value if value is not None else ""
                
                data.append(row_data)
                row_idx += 1
            
            return ExcelReadResult(
                success=True,
                data=data,
                headers=headers,
                row_count=len(data)
            )
            
        except ImportError:
            return ExcelReadResult(
                success=False,
                data=[],
                headers=[],
                row_count=0,
                error_message="openpyxl이 설치되어 있지 않습니다."
            )
        except Exception as e:
            return ExcelReadResult(
                success=False,
                data=[],
                headers=[],
                row_count=0,
                error_message=str(e)
            )
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass
    
    @staticmethod
    def read_excel_streaming(
        file_path: str,
        sheet_name: Optional[str] = None,
        header_row: int = 1,
        chunk_size: int = 100
    ) -> Iterator[list[dict[str, Any]]]:
        """
        대용량 엑셀 파일 스트리밍 읽기 (메모리 최적화)
        
        Args:
            file_path: 엑셀 파일 경로
            sheet_name: 시트 이름
            header_row: 헤더 행 번호
            chunk_size: 청크 크기
        
        Yields:
            데이터 청크 (list[dict])
        """
        from openpyxl import load_workbook
        
        wb = load_workbook(file_path, data_only=True, read_only=True)
        try:
            if sheet_name:
                ws = wb[sheet_name]
            else:
                ws = wb.active

            if ws is None:
                return

            # 헤더 추출
            headers: list[str] = []
            for cell in ws[header_row]:
                value = cell.value
                if value is not None:
                    headers.append(str(value).strip())
                else:
                    headers.append(f"Column_{cell.column}")

            # 데이터 스트리밍
            chunk: list[dict[str, Any]] = []

            for row in ws.iter_rows(min_row=header_row + 1):
                first_cell_value = row[0].value if row else None
                if first_cell_value is None:
                    continue

                row_data: dict[str, Any] = {}
                for idx, cell in enumerate(row):
                    if idx < len(headers):
                        value = cell.value
                        row_data[headers[idx]] = value if value is not None else ""

                chunk.append(row_data)

                if len(chunk) >= chunk_size:
                    yield chunk
                    chunk = []

            # 남은 데이터
            if chunk:
                yield chunk
        finally:
            wb.close()
    
    @staticmethod
    def read_csv(
        file_path: str,
        encoding: str = "utf-8",
        delimiter: str = ","
    ) -> ExcelReadResult:
        """
        CSV 파일 읽기
        
        Args:
            file_path: CSV 파일 경로
            encoding: 파일 인코딩
            delimiter: 구분자
        
        Returns:
            ExcelReadResult 객체
        """
        try:
            path = Path(file_path)
            if not path.exists():
                return ExcelReadResult(
                    success=False,
                    data=[],
                    headers=[],
                    row_count=0,
                    error_message=f"파일이 존재하지 않습니다: {file_path}"
                )
            
            data: list[dict[str, Any]] = []
            headers: list[str] = []
            
            with open(file_path, "r", encoding=encoding, newline="") as f:
                reader = csv.DictReader(f, delimiter=delimiter)
                headers = list(reader.fieldnames or [])
                
                for row in reader:
                    data.append(dict(row))
            
            return ExcelReadResult(
                success=True,
                data=data,
                headers=list(headers),
                row_count=len(data)
            )
            
        except UnicodeDecodeError:
            # CP949로 재시도
            try:
                with open(file_path, "r", encoding="cp949", newline="") as f:
                    reader = csv.DictReader(f, delimiter=delimiter)
                    headers = list(reader.fieldnames or [])
                    data = [dict(row) for row in reader]
                
                return ExcelReadResult(
                    success=True,
                    data=data,
                    headers=list(headers),
                    row_count=len(data)
                )
            except Exception as e:
                return ExcelReadResult(
                    success=False,
                    data=[],
                    headers=[],
                    row_count=0,
                    error_message=f"인코딩 오류: {e}"
                )
        except Exception as e:
            return ExcelReadResult(
                success=False,
                data=[],
                headers=[],
                row_count=0,
                error_message=str(e)
            )
    
    @staticmethod
    def filter_data(
        data: list[dict[str, Any]],
        conditions: dict[str, Any]
    ) -> list[dict[str, Any]]:
        """
        데이터 필터링 (Pandas DataFrame.query 대체)
        
        Args:
            data: 원본 데이터
            conditions: 필터 조건 {컬럼명: 값} 또는 {컬럼명: [값1, 값2]}
        
        Returns:
            필터링된 데이터
        """
        def match_row(row: dict[str, Any]) -> bool:
            for key, value in conditions.items():
                if key not in row:
                    return False
                
                row_value = row[key]
                
                if isinstance(value, (list, tuple)):
                    # IN 조건
                    if row_value not in value:
                        return False
                elif callable(value):
                    # 함수 조건
                    if not value(row_value):
                        return False
                else:
                    # 정확히 일치
                    if row_value != value:
                        return False
            
            return True
        
        return list(filter(match_row, data))
    
    @staticmethod
    def sort_data(
        data: list[dict[str, Any]],
        key: str,
        reverse: bool = False
    ) -> list[dict[str, Any]]:
        """
        데이터 정렬
        
        Args:
            data: 원본 데이터
            key: 정렬 기준 컬럼
            reverse: 내림차순 여부
        
        Returns:
            정렬된 데이터
        """
        return sorted(data, key=lambda x: x.get(key, ""), reverse=reverse)
    
    @staticmethod
    def group_by(
        data: list[dict[str, Any]],
        key: str
    ) -> dict[Any, list[dict[str, Any]]]:
        """
        데이터 그룹화
        
        Args:
            data: 원본 데이터
            key: 그룹화 기준 컬럼
        
        Returns:
            그룹화된 데이터 {그룹값: [행들]}
        """
        groups: dict[Any, list[dict[str, Any]]] = {}
        
        for row in data:
            group_value = row.get(key, "")
            if group_value not in groups:
                groups[group_value] = []
            groups[group_value].append(row)
        
        return groups
    
    @staticmethod
    def select_columns(
        data: list[dict[str, Any]],
        columns: list[str]
    ) -> list[dict[str, Any]]:
        """
        특정 컬럼만 선택
        
        Args:
            data: 원본 데이터
            columns: 선택할 컬럼 목록
        
        Returns:
            선택된 컬럼만 포함된 데이터
        """
        return [
            {k: row.get(k, "") for k in columns}
            for row in data
        ]
    
    @staticmethod
    def write_excel(
        data: list[dict[str, Any]],
        file_path: str,
        sheet_name: str = "Sheet1"
    ) -> bool:
        """
        데이터를 엑셀 파일로 저장
        
        Args:
            data: 저장할 데이터
            file_path: 출력 파일 경로
            sheet_name: 시트 이름
        
        Returns:
            성공 여부
        """
        try:
            from openpyxl import Workbook
            
            if not data:
                return False
            
            wb = Workbook()
            ws = wb.active
            if ws is None:
                return False
            ws.title = sheet_name
            
            # 헤더 작성
            headers = list(data[0].keys())
            for col, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col, value=header)
            
            # 데이터 작성
            for row_idx, row_data in enumerate(data, start=2):
                for col_idx, header in enumerate(headers, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
            
            wb.save(file_path)
            return True
            
        except Exception as e:
            _logger.warning(f"엑셀 저장 실패: {file_path} ({e})", exc_info=True)
            return False
    
    @staticmethod
    def write_csv(
        data: list[dict[str, Any]],
        file_path: str,
        encoding: str = "utf-8-sig"
    ) -> bool:
        """
        데이터를 CSV 파일로 저장
        
        Args:
            data: 저장할 데이터
            file_path: 출력 파일 경로
            encoding: 파일 인코딩
        
        Returns:
            성공 여부
        """
        try:
            if not data:
                return False
            
            headers = list(data[0].keys())
            
            with open(file_path, "w", encoding=encoding, newline="") as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                writer.writerows(data)
            
            return True
            
        except Exception as e:
            _logger.warning(f"CSV 저장 실패: {file_path} ({e})", exc_info=True)
            return False
